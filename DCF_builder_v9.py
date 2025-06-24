import sys
import time
import os
from datetime import datetime
import subprocess
import argparse

import yfinance as yf

import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
import xgboost as xgb

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import NamedStyle, numbers

import matplotlib.pyplot as plt

# --- Handle numpy compatibility with matplotlib/scikit-learn ---
try:
    import matplotlib
    import sklearn
    if int(np.__version__.split('.')[0]) >= 2:
        print("Warning: You are using numpy >=2.0. If you get errors with matplotlib or scikit-learn, run:")
        print("    pip install 'numpy<2.0'")
except Exception as e:
    print(f"Dependency check warning: {e}")

print("yfinance version:", yf.__version__)

def get_unique_filename(ticker, extension=".xlsx"):
    base_name = f"{ticker}_DCF"
    filename = f"{base_name}{extension}"
    counter = 1
    while os.path.exists(filename):
        filename = f"{base_name}_{counter}{extension}"
        counter += 1
    return filename

def robust_fetch_ticker(ticker, max_retries=5):
    for attempt in range(max_retries):
        try:
            return yf.Ticker(ticker)
        except Exception as e:
            # Handle rate limit generically
            if 'rate limit' in str(e).lower():
                wait = 2 ** attempt
                print(f"Rate limit hit, retrying in {wait}s...")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError(f"Failed to fetch data for {ticker} after {max_retries} attempts.")

def fetch_financial_data(ticker):
    try:
        stock = robust_fetch_ticker(ticker)
        info = stock.info
        cashflow = stock.cashflow
        balance_sheet = stock.balance_sheet
        quarterly_balance_sheet = stock.quarterly_balance_sheet

        data = {
            "companyName": info.get("shortName", ticker),
            "totalRevenue": info.get("totalRevenue", 0),
            "ebitda": info.get("ebitda", 0),
            "sharesOutstanding": info.get("sharesOutstanding", 0),
            "currentPrice": info.get("currentPrice", 0)
        }

        # If currentPrice is 0, try other price fields
        if data["currentPrice"] == 0:
            data["currentPrice"] = info.get("regularMarketPrice", 0)
        if data["currentPrice"] == 0:
            data["currentPrice"] = info.get("bid", 0)
        if data["currentPrice"] == 0:
            data["currentPrice"] = info.get("ask", 0)

        if data["currentPrice"] != 0:
            print(f"Using fetched price for {ticker}: {data['currentPrice']}")
        else:
            print(f"Warning: No valid share price found in info for {ticker}. Defaulting to 0.")

        # Try to get totalRevenue and ebitda from income statement if not found in info
        income_stmt = stock.income_stmt
        if not income_stmt.empty:
            if data["totalRevenue"] == 0:
                if "Total Revenue" in income_stmt.index:
                    data["totalRevenue"] = income_stmt.loc["Total Revenue"].iloc[0]
                elif "Revenue" in income_stmt.index:
                    data["totalRevenue"] = income_stmt.loc["Revenue"].iloc[0]
                else:
                    print(f"Warning: No 'Total Revenue' or 'Revenue' found in income statement for {ticker}")

            if data["ebitda"] == 0:
                if "EBITDA" in income_stmt.index:
                    data["ebitda"] = income_stmt.loc["EBITDA"].iloc[0]
                else:
                    print(f"Warning: No 'EBITDA' found in income statement for {ticker}")
        else:
            print(f"Warning: Income statement empty for {ticker}")

        # Depreciation & Amortization
        if not cashflow.empty:
            if "Depreciation And Amortization" in cashflow.index:
                data["depreciation & amortization"] = cashflow.loc["Depreciation And Amortization"].iloc[0]
            elif "Depreciation" in cashflow.index:
                data["depreciation & amortization"] = cashflow.loc["Depreciation"].iloc[0]
            else:
                data["depreciation & amortization"] = 0
                print(f"Warning: No Depreciation data found in cashflow for {ticker}")
        else:
            data["depreciation & amortization"] = 0
            print(f"Warning: Cashflow statement empty for {ticker}")

        # Capital Expenditures
        if not cashflow.empty:
            if "Capital Expenditures" in cashflow.index:
                data["capitalExpenditures"] = cashflow.loc["Capital Expenditures"].iloc[0]
            elif "Capital Expenditure" in cashflow.index:
                data["capitalExpenditures"] = cashflow.loc["Capital Expenditure"].iloc[0]
            else:
                data["capitalExpenditures"] = 0
                print(f"Warning: No Capital Expenditures data found in cashflow for {ticker}")
        else:
            data["capitalExpenditures"] = 0

        def get_working_capital(bs):
            if bs.empty:
                print("Balance sheet is empty")
                return 0
            if "Working Capital" in bs.index:
                return bs.loc["Working Capital"].iloc[0]
            assets_keys = ["Total Current Assets", "Current Assets"]
            liabilities_keys = ["Total Current Liabilities", "Current Liabilities"]
            assets_key = next((key for key in assets_keys if key in bs.index), None)
            liabilities_key = next((key for key in liabilities_keys if key in bs.index), None)
            if assets_key and liabilities_key:
                current_assets = bs.loc[assets_key].iloc[0]
                current_liabilities = bs.loc[liabilities_key].iloc[0]
                return current_assets - current_liabilities
            return 0

        working_capital = get_working_capital(quarterly_balance_sheet)
        if working_capital == 0:
            working_capital = get_working_capital(balance_sheet)
        if working_capital == 0:
            current_assets = info.get("totalCurrentAssets", 0)
            current_liabilities = info.get("totalCurrentLiabilities", 0)
            working_capital = current_assets - current_liabilities
        data["workingCapital"] = working_capital

        print(f"Fetched data: {data}")
        return data
    except Exception as e:
        print(f"Error fetching data for {ticker}: {e}")
        sys.exit(1)

def fetch_historical_financial_data(ticker, years=5):
    try:
        stock = robust_fetch_ticker(ticker)
        financials = stock.financials
        quarterly_financials = stock.quarterly_financials
        info = stock.info
        prices = stock.history(period=f"{years}y")
        
        historical_data = {
            "years": [], 
            "totalRevenue": [], 
            "ebitda": [], 
            "close": [],
            "quarters": [], 
            "quarterlyRevenue": []
        }
        
        # Annual data
        for i in range(min(years, len(financials.columns))):
            year = financials.columns[i].year
            revenue = financials.loc["Total Revenue"].iloc[i] if "Total Revenue" in financials.index else 0
            try:
                ebitda = financials.loc["Operating Income"].iloc[i] + financials.loc["Depreciation"].iloc[i] if "Operating Income" in financials.index and "Depreciation" in financials.index else 0
            except KeyError:
                ebitda = info.get("ebitda", 0) if i == 0 else 0
            
            if pd.notna(revenue) and revenue != 0:
                historical_data["years"].append(year)
                historical_data["totalRevenue"].append(float(revenue))
                historical_data["ebitda"].append(float(ebitda) if pd.notna(ebitda) else 0)
                year_prices = prices[prices.index.year == year]["Close"]
                historical_data["close"].append(year_prices.mean() if not year_prices.empty else 0)
        
        # Quarterly data for beat frequency
        if not quarterly_financials.empty:
            for col in quarterly_financials.columns:
                year = col.year
                quarter = pd.Timestamp(col).quarter
                revenue = quarterly_financials.loc["Total Revenue", col] if "Total Revenue" in quarterly_financials.index else 0
                if pd.notna(revenue) and revenue != 0:
                    historical_data["quarters"].append(f"{year}-Q{quarter}")
                    historical_data["quarterlyRevenue"].append(float(revenue))
        
        print(f"Historical data for {ticker}: {historical_data}")
        return historical_data
    except Exception as e:
        print(f"Error fetching historical data for {ticker}: {e}")
        return {
            "years": [], 
            "totalRevenue": [], 
            "ebitda": [], 
            "close": [], 
            "quarters": [], 
            "quarterlyRevenue": []
        }

def calculate_beat_frequency(ticker, historical_data, lookback_quarters=12):
    if not historical_data["quarters"] or len(historical_data["quarters"]) < 2:
        print("Insufficient quarterly data for beat frequency. Defaulting to 0.5.")
        return 0.5
    
    df = pd.DataFrame({
        "quarter": historical_data["quarters"],
        "revenue": historical_data["quarterlyRevenue"]
    })
    df["revenue_growth"] = df["revenue"].pct_change()
    
    # Heuristic: assume a "beat" if quarterly revenue growth exceeds 5%
    market_avg_growth = 0.05
    beats = (df["revenue_growth"] > market_avg_growth).astype(int)
    
    recent_beats = beats[-lookback_quarters:] if len(beats) > lookback_quarters else beats
    beat_frequency = recent_beats.mean() if len(recent_beats) > 0 else 0.5
    
    print(f"Calculated beat frequency for {ticker}: {beat_frequency:.2f}")
    return beat_frequency

def calculate_growth_projections(historical_data, financial_data, ticker, forecast_years=5, default_growth=0.05):
    if not historical_data["years"] or len(historical_data["years"]) < 2:
        print("Insufficient historical data. Using default growth rate.")
        current_year = datetime.now().year
        projected_years = list(range(current_year, current_year + forecast_years))
        default_revenue = max(financial_data["totalRevenue"], 1_000_000)
        default_ebitda = max(financial_data["ebitda"], 0)
        return {
            "projected_years": projected_years,
            "projected_revenue": [default_revenue * (1 + default_growth) ** (i + 1) for i in range(forecast_years)],
            "projected_ebitda": [default_ebitda * (1 + default_growth) ** (i + 1) for i in range(forecast_years)],
            "short_term_growth": default_growth
        }

    # Create DataFrame with annual data only
    annual_data = {
        "years": historical_data["years"],
        "totalRevenue": historical_data["totalRevenue"],
        "ebitda": historical_data["ebitda"],
        "close": historical_data["close"]
    }
    df = pd.DataFrame(annual_data)
    df["revenue_growth"] = df["totalRevenue"].pct_change() * 100
    df["ebitda_growth"] = df["ebitda"].pct_change() * 100
    df["lagged_revenue"] = df["totalRevenue"].shift(1)
    df["lagged_ebitda"] = df["ebitda"].shift(1)
    df["lagged_growth"] = df["revenue_growth"].shift(1)
    df = df.dropna()

    if len(df) < 2:
        print("Not enough valid data points. Using default growth.")
        current_year = datetime.now().year
        projected_years = list(range(current_year, current_year + forecast_years))
        default_revenue = max(financial_data["totalRevenue"], 1_000_000)
        default_ebitda = max(financial_data["ebitda"], 0)
        return {
            "projected_years": projected_years,
            "projected_revenue": [default_revenue * (1 + default_growth) ** (i + 1) for i in range(forecast_years)],
            "projected_ebitda": [default_ebitda * (1 + default_growth) ** (i + 1) for i in range(forecast_years)],
            "short_term_growth": default_growth
        }

    # Calculate beat frequency from quarterly data
    beat_frequency = calculate_beat_frequency(ticker, historical_data)

    X = pd.DataFrame({
        "lagged_revenue": df["lagged_revenue"],
        "close": df["close"],
        "lagged_growth": df["lagged_growth"],
        "beat_frequency": beat_frequency
    })
    y_revenue = df["revenue_growth"]
    y_ebitda = df["ebitda_growth"]

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    model_revenue = xgb.XGBRegressor(n_estimators=100, learning_rate=0.05, max_depth=3, reg_lambda=1.0, random_state=42)
    model_ebitda = xgb.XGBRegressor(n_estimators=100, learning_rate=0.05, max_depth=3, reg_lambda=1.0, random_state=42)
    
    model_revenue.fit(X_scaled, y_revenue)
    model_ebitda.fit(X_scaled, y_ebitda)

    current_year = datetime.now().year
    projected_years = list(range(current_year, current_year + forecast_years))
    projected_revenue = [financial_data["totalRevenue"]]
    projected_ebitda = [financial_data["ebitda"]]

    latest_growth = df["revenue_growth"].iloc[-1] if not df["revenue_growth"].empty else 0
    for i in range(forecast_years):
        latest_features = scaler.transform([[projected_revenue[-1], historical_data["close"][0], latest_growth, beat_frequency]])
        revenue_growth = model_revenue.predict(latest_features)[0] / 100
        ebitda_growth = model_ebitda.predict(latest_features)[0] / 100
        if ticker == "TSLA":
            revenue_growth = 0.15
            ebitda_growth = 0.15
        else:
            revenue_growth = min(max(revenue_growth, 0.0), 0.3)
            ebitda_growth = min(max(ebitda_growth, 0.0), 0.3)
        next_revenue = max(projected_revenue[-1] * (1 + revenue_growth), 0)
        next_ebitda = max(projected_ebitda[-1] * (1 + ebitda_growth), 0)
        projected_revenue.append(next_revenue)
        projected_ebitda.append(next_ebitda)
        latest_growth = revenue_growth * 100

    projected_revenue = projected_revenue[1:]
    projected_ebitda = projected_ebitda[1:]

    short_term_growth = 0.15 if ticker == "TSLA" else min(max(model_revenue.predict(latest_features)[0] / 100, 0.0), 0.3)

    projections = {
        "projected_years": projected_years,
        "projected_revenue": projected_revenue,
        "projected_ebitda": projected_ebitda,
        "short_term_growth": short_term_growth
    }
    print(f"ML-based growth projections: {projections}")
    return projections

def update_excel_template(template_filename, updated_filename, financial_data, projections, historical_data, mapping):
    try:
        wb = load_workbook(filename=template_filename, data_only=False)
        ws = wb["template"]
        ws.title = "DCF"

        # Register styles only if not already present
        if "accounting" not in wb.named_styles:
            accounting_style = NamedStyle(name="accounting", number_format='_($* #,##0.00_);_($* (#,##0.00);_($* "-"_);_(@_)')
            wb.add_named_style(accounting_style)
        else:
            accounting_style = wb.named_styles["accounting"]

        if "percentage" not in wb.named_styles:
            percentage_style = NamedStyle(name="percentage", number_format='0.00%')
            wb.add_named_style(percentage_style)
        else:
            percentage_style = wb.named_styles["percentage"]

        for cell_address, map_value in mapping.items():
            if map_value in financial_data:
                original_value = ws[cell_address].value
                new_value = financial_data[map_value]
                if cell_address in ["D7", "D8", "D9", "D10", "D11"]:
                    new_value = float(new_value) / 1_000_000
                if map_value != "companyName":
                    new_value = float(new_value) if new_value else 0
                ws[cell_address].value = new_value
                print(f"Updated {cell_address}: {original_value} -> {new_value}")
            elif isinstance(map_value, str):
                original_value = ws[cell_address].value
                ws[cell_address].value = map_value
                print(f"Updated {cell_address}: {original_value} -> {map_value} (literal)")

        ws["D16"].value = projections["short_term_growth"]
        print(f"Set D16 to ML-predicted growth rate: {projections['short_term_growth']*100:.2f}%")

        ws["F24"] = "=D7"
        print(f"Set F24 to formula: =D7")

        # Update valuation cells
        ws["I6"] = "Intrinsic Value Per Share"
        print(f"Set I6 to: Intrinsic Value Per Share")
        ws["K6"] = "=G78/C83"
        ws["K6"].style = accounting_style
        print(f"Set K6 to formula: =G78/C83 with accounting format")
        ws["I7"] = "Current Share Price"
        print(f"Set I7 to: Current Share Price")
        ws["K7"] = financial_data["currentPrice"]
        ws["K7"].style = accounting_style
        print(f"Set K7 to current share price: {financial_data['currentPrice']:.2f} with accounting format")
        ws["I8"] = "Upside/Downside"
        print(f"Set I8 to: Upside/Downside")
        ws["K8"] = "=(K6-K7)/K7"
        ws["K8"].style = percentage_style
        print(f"Set K8 to formula: =(K6-K7)/K7 with percentage format")
        ws["C83"] = financial_data["sharesOutstanding"] / 1_000_000 if financial_data["sharesOutstanding"] > 0 else 0
        print(f"Set C83 to shares outstanding (in millions): {ws['C83'].value:.2f}")

        revenue_values = [financial_data["totalRevenue"]] + projections["projected_revenue"]
        growth_percentages = []
        for i in range(1, len(revenue_values)):
            old_revenue = revenue_values[i - 1]
            new_revenue = revenue_values[i]
            if old_revenue != 0:
                growth = ((new_revenue - old_revenue) / old_revenue) * 100
            else:
                growth = 0
                print(f"Warning: Base revenue is zero for growth calculation.")
            growth_percentages.append(growth)

        # Save the workbook
        wb.save(updated_filename)
        print(f"Saved updated Excel file as {updated_filename}")

    except Exception as e:
        print(f"Error updating Excel template: {e}")
        sys.exit(1)

def main():
    parser = argparse.ArgumentParser(description="DCF Analysis Tool with ML Projections")
    parser.add_argument("--tickers", type=str, default=None, help="Comma-separated stock tickers (e.g., PLTR,MSFT)")
    parser.add_argument("--years", type=int, default=5, help="Number of forecast years (default: 5)")
    parser.add_argument("--growth", type=float, default=0.05, help="Default growth rate if insufficient data (default: 0.05)")
    args = parser.parse_args()

    tickers_input = args.tickers
    if not tickers_input:
        tickers_input = input("Enter stock ticker(s) separated by commas (e.g., PLTR, MSFT): ").strip().upper()

    if not tickers_input:
        print("Ticker(s) cannot be empty.")
        sys.exit(1)

    tickers = [t.strip() for t in tickers_input.split(',') if t.strip()]

    forecast_years = args.years
    default_growth = args.growth

    template_filename = r"C:\Users\micha\OneDrive\Desktop\Claude Build\Exceldcf\Edit-DCFtemplate.xlsx"
    if not os.path.exists(template_filename):
        print(f"File '{template_filename}' does not exist.")
        sys.exit(1)

    for ticker in tickers:
        print(f"\n--- Processing {ticker} ---")

        updated_filename = get_unique_filename(ticker, ".xlsx")
        pdf_filename = get_unique_filename(ticker, "_report.pdf")

        print(f"Fetching current data for {ticker}...")
        financial_data = fetch_financial_data(ticker)

        print(f"Fetching historical data for {ticker}...")
        historical_data = fetch_historical_financial_data(ticker)

        print("Calculating ML-based growth projections...")
        projections = calculate_growth_projections(historical_data, financial_data, ticker, forecast_years, default_growth)

        mapping = {
            "C1": "companyName",
            "D7": "totalRevenue",
            "D8": "ebitda",
            "D9": "depreciation & amortization",
            "D10": "capitalExpenditures",
            "D11": "workingCapital",
            "C9": "Depreciation and Amortization",
            "C29": "Depreciation and Amortization",
            "C39": "Depreciation and Amortization",
            "D5": datetime.now().strftime("%m/%d/%Y"),
            "H1": "Discounted Cash Flow Valuation"
        }
        date_mapping = {chr(70 + i) + "23": f"12/31/{2024 + i}" for i in range(forecast_years)}
        mapping.update(date_mapping)

        print("Updating Excel template with projections...")
        update_excel_template(template_filename, updated_filename, financial_data, projections, historical_data, mapping)

        try:
            if sys.platform == 'win32':
                os.startfile(updated_filename)
            elif sys.platform == 'darwin':
                subprocess.run(['open', updated_filename])
            elif sys.platform.startswith('linux'):
                subprocess.run(['xdg-open', updated_filename])
            else:
                print(f"Unsupported platform: {sys.platform}. Please open {updated_filename} manually.")
        except Exception as e:
            print(f"Error opening file: {e}")

if __name__ == "__main__":
    main()
# For example:
# ticker = "AAPL"
# template_file = "DCF_template.xlsx"
# output_file = get_unique_filename(ticker)
# mapping = {"D7": "totalRevenue", ...}
# financial_data = fetch_financial_data(ticker)
# historical_data = fetch_historical_financial_data(ticker)
# projections = calculate_growth_projections(historical_data, financial_data, ticker)
# update_excel_template(template_file, output_file, financial_data, projections, historical_data, mapping)
